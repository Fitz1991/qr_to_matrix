"""
QR Code Extractor для NiceLabel
================================
При первом запуске автоматически установит все необходимые библиотеки.
Требования: Python 3.8+ (https://python.org)
"""

import sys
import subprocess
import importlib

# ── Автоустановка зависимостей ──────────────────────────────────────────────
REQUIRED = {
    "fitz":     "pymupdf",
    "cv2":      "opencv-python",
    "pyzbar":   "pyzbar",
    "openpyxl": "openpyxl",
    "PIL":      "Pillow",
    "numpy":    "numpy",
}

def auto_install():
    missing = []
    for module, package in REQUIRED.items():
        try:
            importlib.import_module(module)
        except ImportError:
            missing.append(package)
    if missing:
        print(f"Устанавливаю библиотеки: {', '.join(missing)} ...")
        for pkg in missing:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])
        print("Установка завершена! Перезапустите программу.")
        input("Нажмите Enter для выхода...")
        sys.exit(0)

auto_install()

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import csv

import fitz
import cv2
import numpy as np
try:
    from pyzbar.pyzbar import decode as _pyzbar_decode
    PYZBAR_OK = True
except Exception:
    PYZBAR_OK = False
    _pyzbar_decode = None

def decode(img):
    if PYZBAR_OK and _pyzbar_decode:
        return _pyzbar_decode(img)
    return []
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class QRExtractorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("QR Code Extractor → NiceLabel")
        self.root.geometry("640x540")
        self.root.resizable(False, False)
        self.root.configure(bg="#f1f5f9")
        self.pdf_path   = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.results    = []
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self.root, bg="#1d4ed8", height=64)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📄  QR Code Extractor  →  NiceLabel",
                 font=("Segoe UI", 13, "bold"), fg="white", bg="#1d4ed8").pack(expand=True)

        body = tk.Frame(self.root, bg="#f1f5f9", padx=24, pady=18)
        body.pack(fill="both", expand=True)

        self._section(body, "1.  Выберите PDF файл")
        row1 = tk.Frame(body, bg="#f1f5f9"); row1.pack(fill="x", pady=(4, 12))
        tk.Entry(row1, textvariable=self.pdf_path, state="readonly",
                 font=("Segoe UI", 9), bg="white", relief="solid", bd=1
                 ).pack(side="left", fill="x", expand=True, ipady=5)
        self._btn(row1, "Обзор…", self.choose_pdf, "#1d4ed8").pack(side="left", padx=(6, 0))

        self._section(body, "2.  Папка для сохранения результатов")
        row2 = tk.Frame(body, bg="#f1f5f9"); row2.pack(fill="x", pady=(4, 16))
        tk.Entry(row2, textvariable=self.output_dir, state="readonly",
                 font=("Segoe UI", 9), bg="white", relief="solid", bd=1
                 ).pack(side="left", fill="x", expand=True, ipady=5)
        self._btn(row2, "Обзор…", self.choose_output, "#1d4ed8").pack(side="left", padx=(6, 0))

        self.btn_start = self._btn(body, "▶   Начать извлечение QR-кодов",
                                   self.start_extraction, "#15803d",
                                   font=("Segoe UI", 11, "bold"), pady=10)
        self.btn_start.pack(fill="x", pady=(0, 14))

        tk.Label(body, text="Прогресс:", font=("Segoe UI", 9, "bold"),
                 bg="#f1f5f9", fg="#1e293b").pack(anchor="w")
        self.progress = ttk.Progressbar(body, mode="determinate")
        self.progress.pack(fill="x", pady=(4, 10))

        log_wrap = tk.Frame(body, bg="#0f172a")
        log_wrap.pack(fill="both", expand=True)
        self.log = tk.Text(log_wrap, font=("Consolas", 9), bg="#0f172a",
                           fg="#94a3b8", relief="flat", state="disabled", wrap="word", height=9)
        sb = tk.Scrollbar(log_wrap, command=self.log.yview)
        self.log.configure(yscrollcommand=sb.set)
        self.log.pack(side="left", fill="both", expand=True, padx=8, pady=6)
        sb.pack(side="right", fill="y")
        self.log.tag_config("green",  foreground="#4ade80")
        self.log.tag_config("yellow", foreground="#facc15")
        self.log.tag_config("red",    foreground="#f87171")
        self.log.tag_config("blue",   foreground="#60a5fa")

    @staticmethod
    def _section(parent, text):
        tk.Label(parent, text=text, font=("Segoe UI", 10, "bold"),
                 bg="#f1f5f9", fg="#1e293b").pack(anchor="w")

    @staticmethod
    def _btn(parent, text, cmd, color, **kw):
        return tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                         relief="flat", font=kw.pop("font", ("Segoe UI", 9, "bold")),
                         padx=12, cursor="hand2", **kw)

    def log_msg(self, msg, tag=None):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag or "")
        self.log.see("end")
        self.log.configure(state="disabled")
        self.root.update_idletasks()

    def choose_pdf(self):
        p = filedialog.askopenfilename(filetypes=[("PDF файлы", "*.pdf")])
        if p:
            self.pdf_path.set(p)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(p))

    def choose_output(self):
        p = filedialog.askdirectory()
        if p:
            self.output_dir.set(p)

    def start_extraction(self):
        if not self.pdf_path.get():
            messagebox.showwarning("Внимание", "Выберите PDF файл!"); return
        if not self.output_dir.get():
            messagebox.showwarning("Внимание", "Выберите папку для сохранения!"); return
        self.btn_start.configure(state="disabled", bg="#94a3b8")
        self.results = []
        self.progress["value"] = 0
        threading.Thread(target=self._run_extraction, daemon=True).start()

    def _try_decode_image(self, img_cv):
        """Пробует несколько методов распознавания QR-кода."""
        det = cv2.QRCodeDetector()

        # Метод 1: оригинал
        data, _, _ = det.detectAndDecode(img_cv)
        if data:
            return data

        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)

        # Метод 2: grayscale
        data, _, _ = det.detectAndDecode(gray)
        if data:
            return data

        # Метод 3: увеличенный контраст
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        data, _, _ = det.detectAndDecode(enhanced)
        if data:
            return data

        # Метод 4: бинаризация Otsu
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        data, _, _ = det.detectAndDecode(binary)
        if data:
            return data

        # Метод 5: инвертированное изображение (белый QR на тёмном фоне)
        inverted = cv2.bitwise_not(binary)
        data, _, _ = det.detectAndDecode(inverted)
        if data:
            return data

        # Метод 6: адаптивная бинаризация
        adaptive = cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 11, 2
        )
        data, _, _ = det.detectAndDecode(adaptive)
        if data:
            return data

        # Метод 7: заточка (sharpening)
        kern = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
        sharpened = cv2.filter2D(gray, -1, kern)
        data, _, _ = det.detectAndDecode(sharpened)
        if data:
            return data

        return None

    def _run_extraction(self):
        try:
            pdf   = self.pdf_path.get()
            self.log_msg(f"📂 Файл: {os.path.basename(pdf)}")
            doc   = fitz.open(pdf)
            total = len(doc)
            self.log_msg(f"📄 Страниц: {total}")
            self.log_msg(f"🔍 Используется усиленное распознавание (Честный Знак)...")
            found = 0

            # Пробуем разные масштабы рендеринга
            zoom_levels = [6, 8, 4]  # 432dpi, 576dpi, 288dpi

            for i, page in enumerate(doc):
                pn = i + 1
                page_found = False

                for zoom in zoom_levels:
                    if page_found:
                        break

                    mat = fitz.Matrix(zoom, zoom)
                    pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)

                    # Попытка 1: pyzbar
                    pil_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    decoded = decode(pil_img)
                    if decoded:
                        for obj in decoded:
                            data = obj.data.decode("utf-8", errors="replace")
                            self.results.append({"Страница": pn, "Тип": obj.type, "Данные": data})
                            found += 1
                            self.log_msg(f"  ✅ Стр.{pn} [{obj.type}]: {data[:70]}{'…' if len(data)>70 else ''}", "green")
                        page_found = True
                        break

                    # Попытка 2: OpenCV с несколькими методами обработки
                    img_cv = cv2.imdecode(
                        np.frombuffer(pix.tobytes("png"), np.uint8),
                        cv2.IMREAD_COLOR
                    )
                    data = self._try_decode_image(img_cv)
                    if data:
                        self.results.append({"Страница": pn, "Тип": "QRCODE", "Данные": data})
                        found += 1
                        self.log_msg(f"  ✅ Стр.{pn}: {data[:70]}{'…' if len(data)>70 else ''}", "green")
                        page_found = True
                        break

                if not page_found:
                    self.log_msg(f"  ⚠️  Стр.{pn}: не найден", "yellow")

                self.progress["value"] = pn / total * 100
                self.root.update_idletasks()

            doc.close()
            self.log_msg(f"\n📊 Найдено: {found} QR-кодов из {total} страниц")

            if self.results:
                self._save_results()
            else:
                self.log_msg("❌ QR-коды не обнаружены!", "red")
                messagebox.showwarning("Результат", "QR-коды не найдены в документе.")

        except Exception as e:
            self.log_msg(f"❌ Ошибка: {e}", "red")
            messagebox.showerror("Ошибка", str(e))
        finally:
            self.btn_start.configure(state="normal", bg="#15803d")

    def _save_results(self):
        out  = self.output_dir.get()
        base = os.path.splitext(os.path.basename(self.pdf_path.get()))[0]

        csv_path = os.path.join(out, f"{base}_qr_codes.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["Страница", "Тип", "Данные"])
            w.writeheader(); w.writerows(self.results)
        self.log_msg(f"💾 CSV: {csv_path}", "blue")

        xlsx_path = os.path.join(out, f"{base}_qr_codes.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QR Коды"

        thin   = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        h_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        h_fill = PatternFill("solid", fgColor="1D4ED8")
        h_aln  = Alignment(horizontal="center", vertical="center")

        for col, h in enumerate(["Страница", "Тип штрих-кода", "Данные QR-кода  (для NiceLabel)"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = h_font; c.fill = h_fill; c.alignment = h_aln; c.border = border

        even_fill = PatternFill("solid", fgColor="EFF6FF")
        for ri, item in enumerate(self.results, 2):
            fill = even_fill if ri % 2 == 0 else PatternFill()
            for ci, key in enumerate(["Страница", "Тип", "Данные"], 1):
                c = ws.cell(row=ri, column=ci, value=item[key])
                c.border = border; c.fill = fill
                c.alignment = Alignment(vertical="center", wrap_text=(ci == 3))

        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 65
        ws.row_dimensions[1].height = 22
        wb.save(xlsx_path)
        self.log_msg(f"💾 Excel: {xlsx_path}", "blue")
        self.log_msg(f"\n✨ Готово!  Файлы сохранены в:\n   {out}", "green")

        messagebox.showinfo("Готово!",
            f"Извлечено QR-кодов: {len(self.results)}\n\n"
            f"Сохранены файлы:\n"
            f"  • {os.path.basename(csv_path)}\n"
            f"  • {os.path.basename(xlsx_path)}\n\n"
            f"Папка: {out}")


if __name__ == "__main__":
    root = tk.Tk()
    QRExtractorApp(root)
    root.mainloop()
